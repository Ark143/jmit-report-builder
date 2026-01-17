"""
Export API - Export reports to various formats
"""
import frappe
from frappe.decorators import whitelist
import json
from datetime import datetime

@whitelist()
def export_to_pdf(report_name, data):
	"""
	Export report to PDF format
	"""
	try:
		data = json.loads(data) if isinstance(data, str) else data
		
		# Generate PDF using frappe's PDF generation
		html = generate_html_from_data(report_name, data)
		pdf_data = frappe.utils.pdf_lib.get_pdf(html)
		
		return {
			'success': True,
			'file_type': 'pdf',
			'data': pdf_data
		}
	except Exception as e:
		return {
			'success': False,
			'message': str(e)
		}

@whitelist()
def export_to_excel(report_name, data):
	"""
	Export report to Excel format
	"""
	try:
		data = json.loads(data) if isinstance(data, str) else data
		
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment
		
		wb = Workbook()
		ws = wb.active
		ws.title = "Report"
		
		# Add report title
		ws['A1'] = report_name
		ws['A1'].font = Font(bold=True, size=14)
		ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
		
		# Write data
		row = 3
		if isinstance(data, list) and len(data) > 0:
			# Write headers
			if isinstance(data[0], dict):
				headers = list(data[0].keys())
				for col, header in enumerate(headers, 1):
					cell = ws.cell(row=row, column=col, value=header)
					cell.font = Font(bold=True)
					cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
				
				row += 1
				
				# Write data rows
				for record in data:
					for col, header in enumerate(headers, 1):
						ws.cell(row=row, column=col, value=record.get(header, ''))
					row += 1
		
		# Auto-adjust column widths
		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			ws.column_dimensions[column_letter].width = adjusted_width
		
		# Save to bytes
		from io import BytesIO
		output = BytesIO()
		wb.save(output)
		output.seek(0)
		
		return {
			'success': True,
			'file_type': 'xlsx',
			'filename': f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		}
	except Exception as e:
		return {
			'success': False,
			'message': str(e)
		}

@whitelist()
def export_to_csv(report_name, data):
	"""
	Export report to CSV format
	"""
	try:
		data = json.loads(data) if isinstance(data, str) else data
		
		import csv
		from io import StringIO
		
		output = StringIO()
		
		if isinstance(data, list) and len(data) > 0:
			if isinstance(data[0], dict):
				fieldnames = list(data[0].keys())
				writer = csv.DictWriter(output, fieldnames=fieldnames)
				
				writer.writeheader()
				writer.writerows(data)
		
		csv_content = output.getvalue()
		
		return {
			'success': True,
			'file_type': 'csv',
			'filename': f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
			'content': csv_content
		}
	except Exception as e:
		return {
			'success': False,
			'message': str(e)
		}

def generate_html_from_data(report_name, data):
	"""
	Generate HTML representation of report data
	"""
	html = f"""
	<html>
	<head>
		<style>
			body {{ font-family: Arial, sans-serif; margin: 20px; }}
			h1 {{ color: #333; border-bottom: 2px solid #333; padding-bottom: 10px; }}
			table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
			th {{ background-color: #4472C4; color: white; padding: 10px; text-align: left; border: 1px solid #999; }}
			td {{ padding: 8px; border: 1px solid #ddd; }}
			tr:nth-child(even) {{ background-color: #f2f2f2; }}
			.subtotal {{ background-color: #e8e8e8; font-weight: bold; }}
			.group-header {{ background-color: #d9e1f2; font-weight: bold; }}
		</style>
	</head>
	<body>
		<h1>{report_name}</h1>
		<table>
	"""
	
	if isinstance(data, list) and len(data) > 0:
		# Write headers
		if isinstance(data[0], dict):
			headers = list(data[0].keys())
			html += "<tr>"
			for header in headers:
				html += f"<th>{header}</th>"
			html += "</tr>"
			
			# Write data rows
			for record in data:
				row_class = ""
				if record.get('_type') == 'SUBTOTAL':
					row_class = 'class="subtotal"'
				elif record.get('_type') == 'GROUP_HEADER':
					row_class = 'class="group-header"'
				
				html += f"<tr {row_class}>"
				for header in headers:
					html += f"<td>{record.get(header, '')}</td>"
				html += "</tr>"
	
	html += """
		</table>
	</body>
	</html>
	"""
	
	return html
